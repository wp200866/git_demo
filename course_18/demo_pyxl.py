import openpyxl

# 获取工作簿
book = openpyxl.load_workbook('params.xlsx')

print(book.active.cell(column=3, row=3).value)

# 读取工作表
# sheet = book.active
# #
# # 读取单个单元格
# cell_a1 = sheet['A1'].value
# print(cell_a1)
# cell_a3 = sheet.cell(column=1, row=3).value
# print(cell_a3)
#
# # 读取多个连续单元格
# cells = sheet["A1":"C3"]
# print(type(cells), cells)
#
# # 获取单元格值
# cell_a1.value
